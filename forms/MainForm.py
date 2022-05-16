from PyQt5 import QtWidgets, QtCore
from selenium.webdriver.common.by import By
from PyQt5.QtCore import QThread
from PyQt5.QtGui import QCloseEvent
import time
import sys, os
from ui.MainFormACAM import Ui_MainWindow

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from libs.ConnectionDB import ConnectionDB
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.chrome.options import Options
from libs.send_to_mail import SendMail


class MainForm(QtWidgets.QMainWindow):
    def __init__(self, my_window):
        super(MainForm, self).__init__()
        self.my_window = my_window
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowFlags(QtCore.Qt.Window)
        
        self.ui.pushButton.clicked.connect(self.start)
        self.ui.pushButton_2.clicked.connect(self.stop)
        self.ui.pushButton_2.setEnabled(False)
        self.myThread = DownloadExcel(my_window = self)
                     
    def start(self):        
        self.myThread.start()
        self.ui.pushButton.setEnabled(False)
        self.ui.pushButton_2.setEnabled(True)
   
    def forThread(self):
        if self.my_window.first_connection == False:
            self.my_window.driver = webdriver.Chrome()
            self.my_window.driver.maximize_window()
            self.my_window.driver.implicitly_wait(6)
            self.my_window.driver.get("http://ShilkinYuV:!29Ofebov@@sm-sue.fsfk.local/sd/operator/?anchor=")
            time.sleep(1)
            # найти поле имени пользователя / электронной почты и отправить само имя пользователя в поле ввода
            self.my_window.driver.find_element(By.NAME, value="username").send_keys(self.my_window.username)
            # найти поле ввода пароля и также вставить пароль
            self.my_window.driver.find_element(By.NAME, value="password").send_keys(self.my_window.password)
            # нажмите кнопку входа в систему
            self.my_window.driver.find_element(By.CLASS_NAME, value="submit-button").click()

            self.my_window.driver.add_cookie({'name': 'ShilkinYuV', 'value': '!29Ofebov@', 'path': '/'})
            time.sleep(2)
            
            # ждем завершения состояния готовности
            WebDriverWait(driver=self.my_window.driver, timeout=10).until(
                lambda x: x.execute_script("return document.readyState === 'complete'")
            )
            error_message = "Incorrect username or password."
            # получаем ошибки (если есть)
            errors = self.my_window.driver.find_elements(By.CLASS_NAME, value="flash-error")
            # при необходимости распечатать ошибки
            # для e в ошибках:
            #     print(e.text)
            # если мы находим это сообщение об ошибке в составе error, значит вход не выполнен
            if any(error_message in e.text for e in errors):
                print("[!] Login failed")
            else:
                print("[+] Login successful")
                self.my_window.first_connection = True
                # self.forThread()

        self.my_window.driver.get("http://sm-sue.fsfk.local/sd/operator/#uuid:employee$49321051!%7B%22tab%22:%22b661b970-470d-404c-14f6-1144fae3f071,6d426ec2-fb00-edd9-7519-697065c94cf9%22%7D")
        time.sleep(5)
        self.my_window.driver.find_element(By.ID, value="gwt-debug-AdvlistPrsSelectTool.f544db61-174b-675b-008b-00005c7fef0d").click()
        time.sleep(5)     
        self.my_window.driver.find_element(By.XPATH, value=self.my_window.filter1).click()
        print('1')
        time.sleep(5)
        self.my_window.driver.find_element(By.ID, value="gwt-debug-exportAdvlist.f83c557b-1703-ca38-0386-00007da9d7e6").click()
        time.sleep(10)
        self.my_window.driver.find_element(By.ID, value="gwt-debug-AdvlistPrsSelectTool.f544db61-174b-675b-008b-00005c7fef0d").click()
        time.sleep(5)
        self.my_window.driver.find_element(By.XPATH, value=self.my_window.filter2).click()
        print('2')
        time.sleep(6)
        self.my_window.driver.find_element(By.ID, value="gwt-debug-exportAdvlist.f83c557b-1703-ca38-0386-00007da9d7e6").click()
        time.sleep(10)
        self.my_window.driver.find_element(By.ID, value="gwt-debug-AdvlistPrsSelectTool.f544db61-174b-675b-008b-00005c7fef0d").click()
        time.sleep(5)
        self.my_window.driver.find_element(By.XPATH, value=self.my_window.filter3).click()
        print('3')
        time.sleep(6)
        self.my_window.driver.find_element(By.ID, value="gwt-debug-exportAdvlist.f83c557b-1703-ca38-0386-00007da9d7e6").click()
        time.sleep(10)
        self.my_window.first_connection = False
        # self.my_window.driver.close()

    def stop(self):
        self.ui.pushButton.setEnabled(True)
        self.ui.pushButton_2.setEnabled(False)
        self.myThread.terminate()

    def readExcel(self):
        files = os.listdir('C:\\Users\\smsopero\\Downloads')
        realFiles = []
        for file in files:
            if file.__contains__('exportSD'):
                realFiles.append('C:\\Users\\smsopero\\Downloads' + '\\' + file)
        for file in realFiles:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            max_row = ws.max_row
            max_column = ws.max_column

            for i in range(2, max_row + 1):
                registration_date = str(ws.cell(row=i, column=1).value)
                request_type = ws.cell(row=i, column=2).value
                number = ws.cell(row=i, column=3).value
                technical_status = ws.cell(row=i, column=4).value
                service = ws.cell(row=i, column=5).value
                component = ws.cell(row=i, column=6).value
                request_type_two = ws.cell(row=i, column=7).value
                description = ws.cell(row=i, column=8).value
                responsible_group = ws.cell(row=i, column=10).value
                responsible_officer = ws.cell(row=i, column=11).value
                planned_run_time = str(ws.cell(row=i, column=12).value)
                expired_deadline = ws.cell(row=i, column=13).value
                actual_execution_time = ws.cell(row=i, column=14).value
                closing_code = ws.cell(row=i, column=15).value
                solution_description = ws.cell(row=i, column=16).value
                who_decided_the_group = ws.cell(row=i, column=17).value
                who_decided_the_employee = ws.cell(row=i, column=18).value
                location = ws.cell(row=i, column=26).value
                service_recipient = ws.cell(row=i, column=27).value

                connectionDB = ConnectionDB()

                res_number = connectionDB.select_one_number(number)

                if res_number != None:
                    connectionDB.update_one_number(registration_date, request_type, technical_status, service, component, description,
                     responsible_group, responsible_officer, planned_run_time, expired_deadline, actual_execution_time, closing_code,
                      solution_description, who_decided_the_group, who_decided_the_employee, location, service_recipient, request_type_two, number)
                else:
                    connectionDB.insert_all(registration_date, request_type, number, technical_status, service, component, description,
                     responsible_group, responsible_officer, planned_run_time, expired_deadline, actual_execution_time, closing_code,
                      solution_description, who_decided_the_group, who_decided_the_employee, location, service_recipient, request_type_two)

                    mail = SendMail()
                    mail.send_email(number, request_type, registration_date, service, component, service_recipient,location, description, self.my_window.mail_sender, self.my_window.pass_sender, self.my_window.mail_group_one, self.my_window.mail_group_two)


            wb.close()
            os.remove(file)


class DownloadExcel(QThread):

    def __init__(self, my_window):
        QThread.__init__(self)
        super(DownloadExcel, self).__init__()
        self.my_window = my_window

    def run(self):
        while(True):
            try:
                self.my_window.forThread()
            except Exception:
                time.sleep(10)
                self.my_window.forThread()    
            self.my_window.readExcel()
            self.sleep(3600)
            self.my_window.my_window.driver.close()

